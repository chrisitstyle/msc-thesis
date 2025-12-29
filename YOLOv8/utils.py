import csv
import time
from pathlib import Path

import torch

import wandb


# =========================
# global helpers
# =========================

def _append_epoch_stats(save_dir: str, epoch_idx1: int, dt_seconds: float, gpu_mem_gb: float):
    """Appends epoch statistics to a separate epoch_stats.csv file (does not touch results.csv)"""
    path = Path(save_dir) / "epoch_stats.csv"
    is_new = not path.exists()
    with path.open("a", newline="") as f:
        w = csv.writer(f)
        if is_new:
            w.writerow(["epoch", "epoch_time_sec", "epoch_time_min", "gpu_mem_gb"])
        w.writerow([epoch_idx1, f"{dt_seconds:.6f}", f"{dt_seconds / 60.0:.6f}", f"{gpu_mem_gb:.6f}"])


def _integrate_epoch_stats_into_results(save_dir: str):
    """
    After training finishes: read results.csv and epoch_stats.csv, then insert into results.csv:
- 'epoch_time_sec' as the 2nd column,
- 'epoch_time_min' as the 3rd column,
- 'gpu_mem_gb' as the 5th column (i.e., right after the original 2nd column)
    """
    results_path = Path(save_dir) / "results.csv"
    stats_path = Path(save_dir) / "epoch_stats.csv"
    if not results_path.exists() or not stats_path.exists():
        return

    with results_path.open("r", newline="") as f:
        res_rows = list(csv.reader(f))
    with stats_path.open("r", newline="") as f:
        stats_rows = list(csv.reader(f))

    if not res_rows or not stats_rows:
        return

    sh = stats_rows[0]
    try:
        s_epoch = sh.index("epoch")
        s_sec = sh.index("epoch_time_sec")
        s_min = sh.index("epoch_time_min")
        s_gpu = sh.index("gpu_mem_gb")
    except ValueError:
        return

    # Mapping: epoch (1-based) -> (sec, min, gpu)
    stats_map = {}
    for r in stats_rows[1:]:
        if len(r) <= max(s_epoch, s_sec, s_min, s_gpu):
            continue
        try:
            e = int(float(r[s_epoch]))  # epochs are 1-based
        except ValueError:
            continue
        stats_map[e] = (r[s_sec], r[s_min], r[s_gpu])

    header = res_rows[0]
    # New header: [col0] + [sec, min] + [orig col1] + [gpu] + [rest]
    new_header = (
            header[:1]
            + ["epoch_time_sec", "epoch_time_min"]
            + header[1:2]
            + ["gpu_mem_gb"]
            + header[2:]
    )

    new_rows = [new_header]
    for i in range(1, len(res_rows)):
        row = res_rows[i]
        epoch_num = i  # 1-based
        sec, minu, gpu = stats_map.get(epoch_num, ("", "", ""))

        # Build the row in the same order as the header
        new_row = (
                row[:1]  # 1st column (unchanged)
                + [sec, minu]  # 2-3: epoch times
                + row[1:2]  # 4: original 2nd column
                + [gpu]  # 5: gpu_mem_gb
                + row[2:]  # rest of the original columns
        )
        new_rows.append(new_row)

    with results_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerows(new_rows)


def _write_avg_stats_csv(save_dir: str):
    """
    Computes averages from epoch_stats.csv and writes them to epoch_stats_avg.csv
(columns: epoch_time_sec_mean, gpu_mem_gb_mean, epochs_count).
Returns a tuple (mean_sec, mean_gpu, n) or None.
    """
    stats_path = Path(save_dir) / "epoch_stats.csv"
    out_path = Path(save_dir) / "epoch_stats_avg.csv"
    if not stats_path.exists():
        return None

    with stats_path.open("r", newline="") as f:
        rows = list(csv.reader(f))
    if not rows or len(rows) < 2:
        # no epoch data
        with out_path.open("w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["epoch_time_sec_mean", "gpu_mem_gb_mean", "epochs_count"])
            w.writerow(["", "", 0])
        return None

    header = rows[0]
    try:
        sec_idx = header.index("epoch_time_sec")
        gpu_idx = header.index("gpu_mem_gb")
    except ValueError:
        return None

    secs, gpus = [], []
    for r in rows[1:]:
        if len(r) > sec_idx:
            try:
                secs.append(float(r[sec_idx]))
            except Exception:
                pass
        if len(r) > gpu_idx:
            try:
                gpus.append(float(r[gpu_idx]))
            except Exception:
                pass

    n = max(len(secs), len(gpus))
    mean_sec = sum(secs) / len(secs) if secs else 0.0
    mean_gpu = sum(gpus) / len(gpus) if gpus else 0.0

    with out_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["epoch_time_sec_mean", "gpu_mem_gb_mean", "epochs_count"])
        w.writerow([f"{mean_sec:.6f}", f"{mean_gpu:.6f}", n])

    return mean_sec, mean_gpu, n


def on_train_epoch_start(trainer):
    trainer._epoch_tic = time.perf_counter()
    if torch.cuda.is_available():
        torch.cuda.reset_peak_memory_stats()


def on_train_epoch_end(trainer):
    start = getattr(trainer, "_epoch_tic", None)
    if start is None:
        return
    dt = time.perf_counter() - start
    trainer._last_epoch_dt = dt

    if torch.cuda.is_available():
        try:
            torch.cuda.synchronize()
        except Exception:
            pass
        gpu_mem_gb = float(torch.cuda.memory_reserved()) / 1e9
    else:
        gpu_mem_gb = 0.0

    epoch_idx = int(getattr(trainer, "epoch", 0)) + 1

    try:
        wandb.log({
            "epoch": epoch_idx,
            "epoch_time_sec": dt,
            "epoch_time_min": dt / 60.0,
            "gpu_mem_gb": gpu_mem_gb
        })
    except Exception:
        pass

    _append_epoch_stats(trainer.save_dir, epoch_idx, dt, gpu_mem_gb)


def on_train_end(trainer):
    # after Ultralytics closes results.csv
    _integrate_epoch_stats_into_results(trainer.save_dir)
    # compute averages and save to epoch_stats_avg.csv
    avg = _write_avg_stats_csv(trainer.save_dir)
    # optionally log to W&B
    if avg is not None:
        mean_sec, mean_gpu, n = avg
        try:
            wandb.log({
                "avg/epoch_time_sec": mean_sec,
                "avg/gpu_mem_gb": mean_gpu,
                "epochs": n
            })
        except Exception:
            pass

    # XLSX with guaranteed "Text" format (recommended for Excel)
    try:
        _write_results_xlsx(trainer.save_dir)
    except Exception:
        pass


def _write_formatted_results_copy(save_dir: str,
                                  out_name: str = "results_formated.csv",
                                  delimiter: str = ";",
                                  add_excel_sep_hint: bool = True,
                                  csv_force_text: str | None = None):
    """
    Creates a copy of results.csv as results_formated.csv.
- Does not change column order or decimal points.
- add_excel_sep_hint: adds the first line 'sep=;' (Excel will detect the separator).
- csv_force_text: None (by default, do not force) or "formula" => each cell is saved
as =\"...\" (Excel reads it as text without leading apostrophes).
    """
    src_path = Path(save_dir) / "results.csv"
    dst_path = Path(save_dir) / out_name
    if not src_path.exists():
        return

    # Read results.csv with delimiter auto-detection
    with src_path.open("r", newline="") as f:
        sample = f.read(2048)
        f.seek(0)
        try:
            sniffed = csv.Sniffer().sniff(sample)
        except Exception:
            sniffed = csv.excel
            sniffed.delimiter = ","
        rows = list(csv.reader(f, dialect=sniffed))

    with dst_path.open("w", newline="") as f:
        if add_excel_sep_hint and delimiter:
            f.write(f"sep={delimiter}\n")
        w = csv.writer(f, delimiter=delimiter, quoting=csv.QUOTE_ALL)

        if csv_force_text == "formula":
            def cast(v: str) -> str:
                # Save as formula ="..." (escape double quotes)
                return '="' + (v or "").replace('"', '""') + '"'
        else:
            def cast(v: str) -> str:
                return v

        for r in rows:
            w.writerow([cast(x) for x in r])


def _write_results_xlsx(save_dir: str,
                        out_name: str = "results_formated.xlsx",
                        sheet_title: str = "results",
                        freeze_header: bool = False,  # default: no freeze -> no bottom line
                        autosize_columns: bool = True,
                        show_gridlines: bool = True):  # set to False if you want zero grid "borders"
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
    except Exception:
        return

    src_path = Path(save_dir) / "results.csv"
    if not src_path.exists():
        return

    with src_path.open("r", newline="") as f:
        sample = f.read(2048);
        f.seek(0)
        try:
            sniffed = csv.Sniffer().sniff(sample)
        except Exception:
            sniffed = csv.excel;
            sniffed.delimiter = ","
        reader = csv.reader(f, dialect=sniffed)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]
        ws.sheet_view.showGridLines = show_gridlines  # view-only, not borders

        col_widths = []
        header_font = Font(bold=True)

        for ri, row in enumerate(reader, start=1):
            ws.append([("" if v is None else str(v)) for v in row])
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci)
                cell.number_format = "@"  # everything as text
                if ri == 1:
                    cell.font = header_font
                if autosize_columns:
                    ln = len("" if val is None else str(val))
                    if ci - 1 >= len(col_widths):
                        col_widths.extend([0] * (ci - len(col_widths)))
                    col_widths[ci - 1] = max(col_widths[ci - 1], ln)

        if freeze_header:
            ws.freeze_panes = "A2"

        if autosize_columns and col_widths:
            for idx, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = max(w + 2, 1)

        wb.save(Path(save_dir) / out_name)
